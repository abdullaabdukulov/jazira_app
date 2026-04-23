from typing import Dict, List, Optional
from dataclasses import dataclass
from datetime import datetime, date

import frappe
from frappe import _

from jazira_app.jazira_app.utils.helpers import parse_numeric, get_file_path


@dataclass
class ExcelColumn:
    """Excel column configuration."""
    field_name: str
    headers: List[str]


class ExcelService:
    """
    Service for reading and parsing Excel files.

    Supports Uzbek POS report format with columns:
    - Nomi (mahsulot nomi)
    - Soni (miqdori)
    - Narxi (sotuv narxi)
    - Sana (ixtiyoriy - savdo sanasi)
    """

    # Column mapping configuration
    COLUMNS = [
        ExcelColumn("item_name", ["nomi", "mahsulot nomi", "mahsulot", "tovar nomi", "tovar", "наименование"]),
        ExcelColumn("qty", ["soni", "miqdori", "miqdor", "количество", "количество, шт.", "кол-во"]),
        ExcelColumn("rate", ["narxi", "sotuv narxi", "narx", "цена продажи", "цена"]),
        ExcelColumn("datetime", ["sana", "savdo sanasi", "дата", "дата продажи", "datetime", "date"]),
    ]

    # Rows to skip (summary rows)
    SKIP_KEYWORDS = ["jami", "итого", "всего", "total", "сумма", "umumiy"]
    
    def __init__(self):
        self._ensure_openpyxl()
    
    def _ensure_openpyxl(self):
        """Ensure openpyxl is installed."""
        try:
            import openpyxl  # noqa
        except ImportError:
            frappe.throw(_("openpyxl not installed. Run: pip install openpyxl"))
    
    def read_sales_report(self, file_url: str) -> Dict:
        """
        Read POS sales report from Excel file.

        Args:
            file_url: Frappe file URL

        Returns:
            Dict with keys:
                - items: List of dicts with keys: item_name, qty, rate, row_num, date
                - posting_date: str (YYYY-MM-DD) from first row as fallback

        Raises:
            frappe.ValidationError: If file cannot be read or required columns missing
        """
        from openpyxl import load_workbook

        file_path = get_file_path(file_url)
        if not file_path:
            frappe.throw(_("Excel fayl topilmadi: {0}").format(file_url))

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        try:
            # Find header row and column indices
            column_indices = self._find_columns(ws)
            header_row = column_indices.get("_header_row", 1)

            # Validate required columns
            self._validate_required_columns(column_indices)

            # Read data rows
            items = self._read_data_rows(ws, column_indices, header_row)

            # Extract first valid date as fallback/summary date
            excel_date = None
            for item in items:
                if item.get("date"):
                    excel_date = item["date"]
                    break

            return {
                "items": items,
                "posting_date": excel_date
            }

        finally:
            wb.close()
    
    def _find_columns(self, worksheet) -> Dict[str, int]:
        """
        Find column indices by matching headers.
        
        Returns:
            Dict with field_name -> column_index mapping
        """
        column_indices = {}
        
        # Scan first 10 rows for headers
        for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10), start=1):
            for col_num, cell in enumerate(row, start=1):
                if not cell.value:
                    continue
                
                cell_val = str(cell.value).lower().strip()
                
                for column in self.COLUMNS:
                    for header in column.headers:
                        if header in cell_val and column.field_name not in column_indices:
                            column_indices[column.field_name] = col_num
                            column_indices["_header_row"] = row_num
                            break
            
            # Stop if we found required columns
            if "item_name" in column_indices and "qty" in column_indices:
                break
        
        return column_indices
    
    def _validate_required_columns(self, column_indices: Dict):
        """Validate that required columns are found."""
        if "item_name" not in column_indices:
            frappe.throw(_("Excel faylida 'Nomi' ustuni topilmadi"))

        if "qty" not in column_indices:
            frappe.throw(_("Excel faylida 'Soni' ustuni topilmadi"))
    
    def _read_data_rows(
        self,
        worksheet,
        column_indices: Dict[str, int],
        header_row: int
    ) -> List[Dict]:
        """Read and parse data rows."""
        items = []
        
        item_name_col = column_indices["item_name"] - 1
        qty_col = column_indices["qty"] - 1
        rate_col = column_indices.get("rate", 0) - 1 if "rate" in column_indices else None
        dt_col = column_indices.get("datetime", 0) - 1 if "datetime" in column_indices else None
        
        for row_num, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1),
            start=header_row + 1
        ):
            # Get item name
            item_name = self._get_cell_value(row, item_name_col)
            if not item_name:
                continue
            
            # Skip summary rows
            if self._is_summary_row(item_name):
                continue
            
            # Get quantity
            qty = parse_numeric(self._get_cell_value(row, qty_col))
            if qty <= 0:
                continue
            
            # Get rate (optional)
            rate = 0.0
            if rate_col is not None and rate_col >= 0:
                rate = parse_numeric(self._get_cell_value(row, rate_col))

            # Get date (optional per row)
            item_date = None
            if dt_col is not None and dt_col >= 0:
                item_date = self._parse_cell_date(row[dt_col].value)
            
            items.append({
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "row_num": row_num,
                "date": item_date
            })
        
        return items
    
    def _parse_cell_date(self, cell_value) -> Optional[str]:
        """Parse a single cell value into YYYY-MM-DD string."""
        if cell_value is None:
            return None

        # Handle datetime/date objects from Excel
        if isinstance(cell_value, datetime):
            return cell_value.strftime("%Y-%m-%d")
        if isinstance(cell_value, date):
            return cell_value.strftime("%Y-%m-%d")

        # Handle string dates
        cell_str = str(cell_value).strip()
        if not cell_str:
            return None

        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
            try:
                return datetime.strptime(cell_str, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

        return None

    def _get_cell_value(self, row, col_index: int) -> Optional[str]:
        """Safely get cell value."""
        if col_index < 0 or col_index >= len(row):
            return None
        
        value = row[col_index].value
        if value is None:
            return None
        
        return str(value).strip()
    
    def _is_summary_row(self, item_name: str) -> bool:
        """Check if row is a summary/total row."""
        item_lower = item_name.lower()
        return any(kw in item_lower for kw in self.SKIP_KEYWORDS)


excel_service = ExcelService()
