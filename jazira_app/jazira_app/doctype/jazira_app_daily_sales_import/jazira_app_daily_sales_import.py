# -*- coding: utf-8 -*-
# Copyright (c) 2024, Jazira App
# License: MIT
# Jazira App Daily Sales Import - ERPNext v15
# Version 2.1 - VARIANT A: Single Warehouse (Restoran)

"""
WORKFLOW (Bitta Ombor):
1. Excel import
2. BOM li items → Manufacture Stock Entry
   - Raw materials: source_warehouse (-)
   - Finished goods: source_warehouse (+) ← AYNAN SHU OMBORGA!
3. Sales Invoice (Update Stock = ON)
   - All items: source_warehouse (-)
4. BOMsiz items → Direct sale (faqat Sales Invoice)
"""

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt, cint, nowdate, getdate
import hashlib
from typing import Dict, List, Optional, Any


class JaziraAppDailySalesImport(Document):
    """Restaurant Daily Sales Import - Single Warehouse Workflow."""
    
    def validate(self):
        self.validate_warehouse_company()
        
    def validate_warehouse_company(self):
        """Ensure warehouse belongs to selected company."""
        if self.source_warehouse and self.company:
            wh_company = frappe.db.get_value("Warehouse", self.source_warehouse, "company")
            if wh_company and wh_company != self.company:
                frappe.throw(
                    _("Warehouse {0} does not belong to Company {1}").format(
                        self.source_warehouse, self.company
                    )
                )
    
    def before_submit(self):
        frappe.throw(_("This document cannot be submitted. Use 'Process Import' button."))


# =============================================================================
# EXCEL READING FUNCTIONS
# =============================================================================

def parse_numeric(value: Any) -> float:
    """Parse numeric value from various formats (European/US)."""
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        return float(value)
    
    str_val = str(value).strip()
    if not str_val:
        return 0.0
    
    # Remove spaces (thousand separators)
    str_val = str_val.replace(" ", "")
    
    # Handle comma/dot formats
    if "," in str_val:
        dots = str_val.count(".")
        commas = str_val.count(",")
        
        if commas == 1 and dots == 0:
            str_val = str_val.replace(",", ".")
        elif commas == 1 and dots >= 1:
            str_val = str_val.replace(".", "").replace(",", ".")
        elif dots == 0 and commas > 1:
            str_val = str_val.replace(",", "")
    
    if "." in str_val and str_val.count(".") == 1:
        parts = str_val.split(".")
        if len(parts[1]) == 3 and parts[1].isdigit():
            str_val = str_val.replace(".", "")
    
    try:
        return float(str_val)
    except ValueError:
        return 0.0


def read_excel_file(file_url: str) -> List[Dict]:
    """Read Excel file and return list of dictionaries."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        frappe.throw(_("openpyxl not installed. Run: pip install openpyxl"))
    
    # Get actual file path
    if file_url.startswith("/private/files/") or file_url.startswith("/files/"):
        file_path = frappe.get_site_path() + file_url
    else:
        file_doc = frappe.db.get_value("File", {"file_url": file_url}, ["file_url"], as_dict=True)
        if file_doc:
            file_path = frappe.get_site_path() + file_doc.file_url
        else:
            file_path = frappe.get_site_path() + file_url
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Column mapping (Russian headers)
    COLUMN_MAP = {
        "наименование": "item_name",
        "количество": "qty",
        "количество, шт.": "qty",
        "количество, шт": "qty",
        "кол-во": "qty",
        "цена продажи": "rate",
        "цена продажи, uzs. коп.": "rate",
        "цена продажи, uzs": "rate",
        "цена": "rate",
    }
    
    header_row = None
    column_indices = {}
    
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
        for col_num, cell in enumerate(row, start=1):
            if cell.value:
                cell_val = str(cell.value).lower().strip()
                for ru_header, field_name in COLUMN_MAP.items():
                    if ru_header in cell_val:
                        column_indices[field_name] = col_num
                        header_row = row_num
        
        if "item_name" in column_indices and "qty" in column_indices:
            break
    
    if not header_row or "item_name" not in column_indices:
        frappe.throw(_("Excel faylida 'Наименование' ustuni topilmadi."))
    
    if "qty" not in column_indices:
        frappe.throw(_("Excel faylida 'количество' ustuni topilmadi."))
    
    items = []
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        item_name_idx = column_indices["item_name"] - 1
        item_name = row[item_name_idx].value if item_name_idx < len(row) else None
        
        if not item_name or str(item_name).strip() == "":
            continue
        
        item_name = str(item_name).strip()
        
        # Skip total/summary rows
        skip_keywords = ["итого", "всего", "total", "сумма", "jami"]
        if any(kw in item_name.lower() for kw in skip_keywords):
            continue
        
        qty_idx = column_indices["qty"] - 1
        qty = parse_numeric(row[qty_idx].value if qty_idx < len(row) else 0)
        
        if qty <= 0:
            continue
        
        rate = 0.0
        if "rate" in column_indices:
            rate_idx = column_indices["rate"] - 1
            rate = parse_numeric(row[rate_idx].value if rate_idx < len(row) else 0)
        
        items.append({
            "item_name": item_name,
            "qty": qty,
            "rate": rate,
            "row_num": row_num
        })
    
    wb.close()
    return items


def calculate_excel_hash(file_url: str) -> str:
    """Calculate MD5 hash of Excel file for duplicate detection."""
    if file_url.startswith("/private/files/") or file_url.startswith("/files/"):
        file_path = frappe.get_site_path() + file_url
    else:
        file_doc = frappe.db.get_value("File", {"file_url": file_url}, "file_url")
        if file_doc:
            file_path = frappe.get_site_path() + file_doc
        else:
            file_path = frappe.get_site_path() + file_url
    
    try:
        with open(file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    except Exception:
        return ""


# =============================================================================
# VALIDATION FUNCTIONS
# =============================================================================

@frappe.whitelist()
def validate_before_upload(company: str, source_warehouse: str, posting_date: str) -> Dict:
    """Validate prerequisites before allowing Excel upload."""
    errors = []
    
    if not company:
        errors.append(_("Company tanlanmagan"))
    
    if not source_warehouse:
        errors.append(_("Oshxona Ombori tanlanmagan"))
    
    if not posting_date:
        errors.append(_("Posting Date (savdo sanasi) tanlanmagan"))
    
    # Validate warehouse belongs to company
    if company and source_warehouse:
        wh_company = frappe.db.get_value("Warehouse", source_warehouse, "company")
        if wh_company and wh_company != company:
            errors.append(_("Warehouse '{0}' kompaniyaga tegishli emas: {1}").format(
                source_warehouse, company
            ))
    
    # Check customer exists
    if not frappe.db.exists("Customer", "Walk-in Customer"):
        errors.append(_("'Walk-in Customer' nomli mijoz topilmadi. Iltimos, avval yarating."))
    
    if errors:
        return {"success": False, "message": "\n".join(errors)}
    
    return {"success": True, "message": _("Barcha tekshiruvlar muvaffaqiyatli o'tdi")}


@frappe.whitelist()
def validate_excel_items(doc_name: str) -> Dict:
    """Validate items in uploaded Excel file."""
    doc = frappe.get_doc("Jazira App Daily Sales Import", doc_name)
    
    if not doc.excel_file:
        return {"success": False, "message": _("Excel fayl yuklanmagan"), "errors": [], "items": []}
    
    try:
        excel_items = read_excel_file(doc.excel_file)
    except Exception as e:
        return {"success": False, "message": _("Excel o'qishda xato: {0}").format(str(e)), "errors": [], "items": []}
    
    if not excel_items:
        return {"success": False, "message": _("Excel faylda hech qanday sotuv topilmadi"), "errors": [], "items": []}
    
    errors = []
    valid_items = []
    
    for item in excel_items:
        item_name = item["item_name"]
        item_code = frappe.db.get_value("Item", {"item_name": item_name}, "name")
        
        if not item_code:
            item_code = frappe.db.get_value("Item", {"item_name": ["like", f"%{item_name}%"]}, "name")
        
        if not item_code:
            errors.append({
                "row": item["row_num"],
                "item_name": item_name,
                "error": _("Item topilmadi: '{0}'").format(item_name)
            })
        else:
            item["item_code"] = item_code
            item["found"] = True
            valid_items.append(item)
    
    # Check duplicate import
    if doc.excel_file:
        excel_hash = calculate_excel_hash(doc.excel_file)
        existing = frappe.db.exists(
            "Jazira App Daily Sales Import",
            {"external_ref": excel_hash, "name": ["!=", doc_name], "status": "Processed"}
        )
        if existing:
            errors.insert(0, {
                "row": 0,
                "item_name": "",
                "error": _("Bu Excel fayl avval import qilingan: {0}").format(existing)
            })
    
    return {
        "success": len(errors) == 0,
        "message": _("{0} ta item topildi, {1} ta xato").format(len(valid_items), len(errors)),
        "errors": errors,
        "items": valid_items,
        "total_qty": sum(i["qty"] for i in valid_items),
        "total_amount": sum(i["qty"] * i["rate"] for i in valid_items)
    }


# =============================================================================
# BOM FUNCTIONS
# =============================================================================

def get_item_bom(item_code: str) -> Optional[str]:
    """Get default active BOM for an item."""
    bom = frappe.db.get_value(
        "BOM",
        {"item": item_code, "is_default": 1, "is_active": 1, "docstatus": 1},
        "name"
    )
    return bom


def get_bom_items(bom_name: str, qty: float) -> List[Dict]:
    """Get BOM items (raw materials) with calculated quantities."""
    from frappe.query_builder import DocType
    
    BOMItem = DocType("BOM Item")
    
    bom_items = (
        frappe.qb.from_(BOMItem)
        .select(BOMItem.item_code, BOMItem.qty, BOMItem.uom, BOMItem.stock_qty, BOMItem.stock_uom)
        .where(BOMItem.parent == bom_name)
        .run(as_dict=True)
    )
    
    bom_qty = frappe.db.get_value("BOM", bom_name, "quantity") or 1
    
    raw_materials = []
    for bi in bom_items:
        required_qty = (bi.stock_qty / bom_qty) * qty
        raw_materials.append({
            "item_code": bi.item_code,
            "qty": required_qty,
            "uom": bi.stock_uom or bi.uom
        })
    
    return raw_materials


# =============================================================================
# MANUFACTURE STOCK ENTRY (VARIANT A - SINGLE WAREHOUSE)
# =============================================================================

def create_manufacture_stock_entries(
    doc: Document,
    items_with_bom: List[Dict],
    submit: bool = True
) -> List[str]:
    """
    Create Manufacture Stock Entries - ONE PER FINISHED ITEM.
    
    ERPNext limitation: Only ONE is_finished_item per Stock Entry!
    So we create separate Stock Entry for each BOM item.
    
    VARIANT A Logic:
    - Raw materials: source_warehouse → consumed (MINUS)
    - Finished goods: source_warehouse → produced (PLUS)
    - BOTH operations in the SAME warehouse!
    """
    if not items_with_bom:
        return []
    
    log_entries = ["=" * 50, "MANUFACTURE STOCK ENTRIES", "=" * 50]
    log_entries.append(f"Warehouse: {doc.source_warehouse} (bitta ombor)")
    log_entries.append(f"BOM li itemlar: {len(items_with_bom)} ta")
    log_entries.append("")
    
    stock_entry_names = []
    
    # Handle negative stock
    if doc.allow_negative_stock:
        frappe.flags.allow_negative_stock = True
    
    try:
        # Create separate Stock Entry for EACH BOM item
        for item_data in items_with_bom:
            item_code = item_data["item_code"]
            qty = item_data["qty"]
            bom = item_data["bom"]
            
            log_entries.append(f"📦 TAOM: {item_data['item_name']} x {qty}")
            log_entries.append(f"   BOM: {bom}")
            
            se = frappe.new_doc("Stock Entry")
            se.stock_entry_type = "Manufacture"
            se.company = doc.company
            se.posting_date = doc.posting_date
            se.posting_time = "23:59:59"
            
            # VARIANT A: Both from and to warehouse = source_warehouse
            se.from_warehouse = doc.source_warehouse
            se.to_warehouse = doc.source_warehouse
            
            # Get raw materials from BOM
            raw_materials = get_bom_items(bom, qty)
            
            # Add raw material rows (consumed from source_warehouse)
            for rm in raw_materials:
                se.append("items", {
                    "item_code": rm["item_code"],
                    "qty": rm["qty"],
                    "uom": rm["uom"],
                    "s_warehouse": doc.source_warehouse,
                    "t_warehouse": None,
                    "is_finished_item": 0,
                    "allow_zero_valuation_rate": 1
                })
                log_entries.append(f"   ➖ {rm['item_code']}: {rm['qty']} {rm['uom']}")
            
            # Add finished item row (ONLY ONE per Stock Entry!)
            item_uom = frappe.db.get_value("Item", item_code, "stock_uom") or "Nos"
            se.append("items", {
                "item_code": item_code,
                "qty": qty,
                "uom": item_uom,
                "s_warehouse": None,
                "t_warehouse": doc.source_warehouse,
                "is_finished_item": 1,
                "bom_no": bom,
                "allow_zero_valuation_rate": 1
            })
            log_entries.append(f"   ➕ {item_code}: {qty} {item_uom} → {doc.source_warehouse}")
            
            se.flags.ignore_permissions = True
            se.insert()
            
            if submit:
                se.submit()
            
            stock_entry_names.append(se.name)
            log_entries.append(f"   ✅ Stock Entry: {se.name}")
            log_entries.append("")
        
        log_entries.append(f"✅ Jami {len(stock_entry_names)} ta Manufacture Stock Entry yaratildi")
        
    except Exception as e:
        log_entries.append(f"❌ Stock Entry xatosi: {str(e)}")
        raise
    finally:
        frappe.flags.allow_negative_stock = False
    
    # Save log
    current_log = doc.import_log or ""
    doc.db_set("import_log", current_log + "\n".join(log_entries) + "\n\n")
    
    return stock_entry_names


# =============================================================================
# SALES INVOICE (VARIANT A - SINGLE WAREHOUSE)
# =============================================================================

def create_sales_invoice(
    doc: Document,
    items: List[Dict],
    submit: bool = True
) -> str:
    """
    Create Sales Invoice with Update Stock ON.
    
    VARIANT A Logic:
    - All items sold from source_warehouse
    - Update Stock = 1 (stock automatically reduced)
    """
    log_entries = ["=" * 50, "SALES INVOICE", "=" * 50]
    log_entries.append(f"Customer: {doc.customer or 'Walk-in Customer'}")
    log_entries.append(f"Warehouse: {doc.source_warehouse}")
    log_entries.append(f"Update Stock: ON")
    log_entries.append("")
    
    si = frappe.new_doc("Sales Invoice")
    si.company = doc.company
    si.customer = doc.customer or "Walk-in Customer"
    si.posting_date = doc.posting_date
    si.posting_time = "23:59:59"
    si.due_date = doc.posting_date
    
    # UPDATE STOCK = ON (stock reduced automatically)
    si.update_stock = 1
    si.set_warehouse = doc.source_warehouse  # ← ALL SALES FROM THIS WAREHOUSE
    
    for item_data in items:
        si.append("items", {
            "item_code": item_data["item_code"],
            "qty": item_data["qty"],
            "rate": item_data["rate"],
            "warehouse": doc.source_warehouse,  # ← MINUS from here
            "allow_zero_valuation_rate": 1  # ← YANGI: Valuation Rate bo'lmasa ham ishlaydi
        })
        log_entries.append(f"📄 {item_data['item_name']}: {item_data['qty']} x {item_data['rate']} = {item_data['qty'] * item_data['rate']}")
    
    si.flags.ignore_permissions = True
    si.insert()
    
    if submit:
        si.submit()
    
    log_entries.append("")
    log_entries.append(f"💰 Jami: {sum(i['qty'] * i['rate'] for i in items)}")
    log_entries.append(f"✅ Sales Invoice yaratildi: {si.name}")
    
    # Save log
    current_log = doc.import_log or ""
    doc.db_set("import_log", current_log + "\n".join(log_entries) + "\n")
    
    return si.name


# =============================================================================
# MAIN PROCESSING FUNCTION
# =============================================================================

@frappe.whitelist()
def process_import(doc_name: str, background: bool = False) -> Dict:
    """Main function to process the import."""
    if background:
        frappe.enqueue(
            _process_import_background,
            queue="long",
            timeout=3600,
            doc_name=doc_name
        )
        return {"success": True, "message": _("Import jarayoni fonada boshlandi")}
    
    return _process_import_sync(doc_name)


def _process_import_background(doc_name: str):
    """Background job wrapper."""
    try:
        result = _process_import_sync(doc_name)
        event = "restaurant_import_success" if result["success"] else "restaurant_import_failed"
        frappe.publish_realtime(
            event,
            {"doc_name": doc_name, "result": result},
            doctype="Jazira App Daily Sales Import",
            docname=doc_name
        )
    except Exception as e:
        frappe.log_error(f"Import Error: {doc_name}\n{str(e)}", "Jazira App Daily Sales Import")
        doc = frappe.get_doc("Jazira App Daily Sales Import", doc_name)
        doc.db_set("status", "Failed")
        doc.db_set("error_log", str(e))


def _process_import_sync(doc_name: str) -> Dict:
    """
    Synchronous import processing - VARIANT A (Single Warehouse).
    
    Workflow:
    1. Validate prerequisites
    2. Read Excel and match items
    3. Separate items: with BOM vs without BOM
    4. Create Manufacture Stock Entry (BOM items only)
       - Raw materials: source_warehouse (-)
       - Finished goods: source_warehouse (+)
    5. Create Sales Invoice (ALL items)
       - All items: source_warehouse (-)
       - Update Stock = ON
    """
    doc = frappe.get_doc("Jazira App Daily Sales Import", doc_name)
    
    # Check idempotency
    if doc.status == "Processed":
        return {
            "success": False,
            "message": _("Bu import allaqachon bajarilgan.\nStock Entry: {0}\nSales Invoice: {1}").format(
                doc.stock_entry or "N/A", doc.sales_invoice
            )
        }
    
    doc.db_set("status", "Processing")
    doc.db_set("error_log", "")
    doc.db_set("import_log", f"{'=' * 50}\nIMPORT BOSHLANDI: {nowdate()}\nVARIANT A: Bitta Ombor\n{'=' * 50}\n\n")
    
    try:
        # 1. Validate
        validation = validate_before_upload(
            doc.company, 
            doc.source_warehouse, 
            str(doc.posting_date)
        )
        if not validation["success"]:
            raise Exception(validation["message"])
        
        # 2. Read Excel
        excel_items = read_excel_file(doc.excel_file)
        if not excel_items:
            raise Exception(_("Excel faylda sotuv topilmadi"))
        
        # 3. Check duplicate
        excel_hash = calculate_excel_hash(doc.excel_file)
        existing = frappe.db.exists(
            "Jazira App Daily Sales Import",
            {"external_ref": excel_hash, "name": ["!=", doc_name], "status": "Processed"}
        )
        if existing:
            raise Exception(_("Bu Excel avval import qilingan: {0}").format(existing))
        
        # 4. Match items and separate by BOM
        errors = []
        items_with_bom = []
        items_without_bom = []
        all_valid_items = []
        
        for item in excel_items:
            item_code = frappe.db.get_value("Item", {"item_name": item["item_name"]}, "name")
            
            if not item_code:
                errors.append(_("Qator {0}: Item topilmadi - '{1}'").format(item["row_num"], item["item_name"]))
                continue
            
            item["item_code"] = item_code
            all_valid_items.append(item)
            
            # Check BOM
            bom = get_item_bom(item_code)
            if bom:
                item["bom"] = bom
                items_with_bom.append(item)
            else:
                items_without_bom.append(item)
        
        if errors:
            raise Exception("\n".join(errors))
        
        # Log summary
        log_summary = [
            f"📊 ITEMS SUMMARY:",
            f"   Jami: {len(all_valid_items)}",
            f"   BOM bilan (Manufacture): {len(items_with_bom)}",
            f"   BOMsiz (Direct Sale): {len(items_without_bom)}",
            ""
        ]
        current_log = doc.import_log or ""
        doc.db_set("import_log", current_log + "\n".join(log_summary) + "\n")
        
        # 5. Create Manufacture Stock Entries (BOM items only) - ONE PER ITEM!
        se_names = []
        if items_with_bom:
            se_names = create_manufacture_stock_entries(doc, items_with_bom, submit=True)
            # Save all SE names as comma-separated list
            if se_names:
                doc.db_set("stock_entry", ", ".join(se_names))
        else:
            current_log = doc.import_log or ""
            doc.db_set("import_log", current_log + "ℹ️ BOM li mahsulot yo'q - Manufacture SKIP qilindi\n\n")
        
        # 6. Create Sales Invoice (ALL items)
        si_name = create_sales_invoice(doc, all_valid_items, submit=True)
        doc.db_set("sales_invoice", si_name)
        
        # 7. Finalize
        doc.db_set("external_ref", excel_hash)
        doc.db_set("status", "Processed")
        
        # Final log
        final_log = [
            "",
            "=" * 50,
            "✅ IMPORT MUVAFFAQIYATLI YAKUNLANDI",
            "=" * 50,
        ]
        current_log = doc.import_log or ""
        doc.db_set("import_log", current_log + "\n".join(final_log))
        
        frappe.db.commit()
        
        return {
            "success": True,
            "message": _("Import muvaffaqiyatli bajarildi"),
            "stock_entries": se_names,
            "sales_invoice": si_name,
            "items_with_bom": len(items_with_bom),
            "items_without_bom": len(items_without_bom),
            "total_items": len(all_valid_items),
            "total_amount": sum(i["qty"] * i["rate"] for i in all_valid_items)
        }
        
    except Exception as e:
        frappe.db.rollback()
        doc.db_set("status", "Failed")
        doc.db_set("error_log", str(e))
        frappe.log_error(f"Import Error: {doc_name}\n{str(e)}", "Jazira App Daily Sales Import")
        return {"success": False, "message": str(e)}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

@frappe.whitelist()
def get_preview_data(doc_name: str) -> Dict:
    """Get preview of Excel data before processing."""
    doc = frappe.get_doc("Jazira App Daily Sales Import", doc_name)
    
    if not doc.excel_file:
        return {"success": False, "message": _("Excel fayl yuklanmagan")}
    
    try:
        items = read_excel_file(doc.excel_file)
        
        for item in items:
            item_code = frappe.db.get_value("Item", {"item_name": item["item_name"]}, "name")
            item["item_code"] = item_code
            item["found"] = bool(item_code)
            
            if item_code:
                bom = get_item_bom(item_code)
                item["has_bom"] = bool(bom)
                item["bom"] = bom
                item["type"] = "MANUFACTURE" if bom else "DIRECT SALE"
            else:
                item["has_bom"] = False
                item["type"] = "NOT FOUND"
        
        return {
            "success": True,
            "items": items,
            "summary": {
                "total_items": len(items),
                "found": len([i for i in items if i["found"]]),
                "not_found": len([i for i in items if not i["found"]]),
                "with_bom": len([i for i in items if i.get("has_bom")]),
                "without_bom": len([i for i in items if i["found"] and not i.get("has_bom")]),
                "total_qty": sum(i["qty"] for i in items),
                "total_amount": sum(i["qty"] * i["rate"] for i in items)
            }
        }
        
    except Exception as e:
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def cancel_import(doc_name: str) -> Dict:
    """Cancel a processed import - cancels related documents."""
    doc = frappe.get_doc("Jazira App Daily Sales Import", doc_name)
    
    if doc.status != "Processed":
        return {"success": False, "message": _("Faqat 'Processed' statusdagi importni bekor qilish mumkin")}
    
    try:
        # Cancel Sales Invoice first (depends on stock)
        if doc.sales_invoice:
            si = frappe.get_doc("Sales Invoice", doc.sales_invoice)
            if si.docstatus == 1:
                si.cancel()
        
        # Cancel ALL Stock Entries (comma-separated list)
        if doc.stock_entry:
            se_names = [se.strip() for se in doc.stock_entry.split(",") if se.strip()]
            for se_name in se_names:
                if frappe.db.exists("Stock Entry", se_name):
                    se = frappe.get_doc("Stock Entry", se_name)
                    if se.docstatus == 1:
                        se.cancel()
        
        doc.db_set("status", "Draft")
        doc.db_set("external_ref", "")
        doc.db_set("import_log", "")
        
        frappe.db.commit()
        
        return {"success": True, "message": _("Import bekor qilindi")}
        
    except Exception as e:
        frappe.db.rollback()
        return {"success": False, "message": str(e)}


@frappe.whitelist()
def get_default_warehouse(company: str) -> Dict:
    """Get default warehouse for a company."""
    if not company:
        return {"source_warehouse": None}
    
    # Try to find default warehouse
    # 1. First try "Stores" type
    source_wh = frappe.db.get_value(
        "Warehouse",
        {"company": company, "is_group": 0, "warehouse_type": "Stores"},
        "name"
    )
    
    # 2. Then try any non-group warehouse
    if not source_wh:
        source_wh = frappe.db.get_value(
            "Warehouse",
            {"company": company, "is_group": 0},
            "name"
        )
    
    return {"source_warehouse": source_wh}